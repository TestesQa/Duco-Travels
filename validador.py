from pickle import APPEND
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
from collections import Counter

#Carrega planilha completa
wb = load_workbook('agendas/Agenda.xlsx')

#Lista as abas dentro da planilha
for sheet in wb:
    print(sheet)

#Seleciona Aba a ser utilizada
ws = wb['teste.xlsx']

#Cria lista 
agenda = []
   
#Preenche a lista com os dados necessário, informar no rangue onde começa e onde termina e informar qual coluna vai extrair os dados.
for n in range(3,10653):

    col = str(n)
    my_value = ws['E'+ col].value
    if my_value != 'Break' or my_value != None:
        agenda.append(my_value)  
        n = int(n)
    else:
        pass
    

#Criar contador
contador = Counter(agenda)

# Cria lista de repetidos

repetidos = [
        item for item, quantidade in contador.items() 
            if quantidade > 1
    ]

#Apresenta valor de repeditos
quantidade_repetidos = len(repetidos)-2


#Printa lista de repetido e quantidade
print("\n \n ""Lista de Nome Repetidos --> ",repetidos,"\n Quantidade de repetidos sem BREAK e NONE  --> ", quantidade_repetidos )
    




